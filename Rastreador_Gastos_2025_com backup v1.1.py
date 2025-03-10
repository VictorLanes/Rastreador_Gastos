import os
import shutil
import sqlite3
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime, timedelta

import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.tooltip import ToolTip
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class SpendingTracker(ttk.Window):
    def __init__(self):
        super().__init__(themename="cosmo")
        self.title("Rastreador de Gastos")
        self.geometry("1100x650")

        # Conectar ao banco de dados e carregar os dados
        self.conectar_banco()
        self.carregar_dados()

        # -------------------------------------------------
        # VARIÁVEIS DE CONTROLE (Despesas)
        # -------------------------------------------------
        # Cada despesa: (Ano, Mês, Despesa, Valor, Vencimento, Categoria, Observação, Pagamento, Cartão Utilizado)
        if not hasattr(self, 'despesas'):
            self.despesas = []
        self.categorias = ["Alimentação", "Transporte", "Lazer", "Saúde", "Moradia", "Outros"]
        self.formas_pagamento = ["VR", "Cartão de Crédito", "PIX", "Débito"]
        self.orcamento_mensal = 0.0  # Apenas despesas via Débito ou PIX diminuem o saldo

        # -------------------------------------------------
        # VARIÁVEIS DE CONTROLE (Cartão de Crédito)
        # -------------------------------------------------
        # Cada cartão: (id, nome_cartao, nome_usuario, número (mascarado), validade, bandeira, limite)
        if not hasattr(self, 'cartoes'):
            self.cartoes = []

        # -------------------------------------------------
        # VARIÁVEIS DE CONTROLE (Metas)
        # -------------------------------------------------
        # Cada meta: (id, nome, valor_meta, valor_atual, data_inicial, data_final)
        self.metas = []

        # -------------------------------------------------
        # NOTEBOOK (Abas)
        # -------------------------------------------------
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=ttk.BOTH, expand=True)

        self.tab_despesas = ttk.Frame(self.notebook, padding=10)
        self.tab_cartao = ttk.Frame(self.notebook, padding=10)
        self.tab_relatorios = ttk.Frame(self.notebook, padding=10)
        self.tab_metas = ttk.Frame(self.notebook, padding=10)

        self.notebook.add(self.tab_despesas, text="Despesas")
        self.notebook.add(self.tab_cartao, text="Cartão de Crédito")
        self.notebook.add(self.tab_relatorios, text="Relatórios")
        self.notebook.add(self.tab_metas, text="Metas")

        # -------------------------------------------------
        # ABA: DESPESAS
        # -------------------------------------------------
        self.configurar_aba_despesas()
        # -------------------------------------------------
        # ABA: CARTÃO DE CRÉDITO
        # -------------------------------------------------
        self.configurar_aba_cartao()
        # -------------------------------------------------
        # ABA: RELATÓRIOS
        # -------------------------------------------------
        self.configurar_aba_relatorios()
        # -------------------------------------------------
        # ABA: METAS
        # -------------------------------------------------
        self.configurar_aba_metas()

    # ----- Conexão com o Banco de Dados (SQLite) -----
    def conectar_banco(self):
        self.conn = sqlite3.connect("financeiro.db")
        self.cursor = self.conn.cursor()
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS despesas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ano TEXT,
                mes TEXT,
                despesa TEXT,
                valor REAL,
                vencimento TEXT,
                categoria TEXT,
                observacao TEXT,
                pagamento TEXT,
                cartao_utilizado TEXT
            )
        """)
        # Tabela Cartões – removidos os campos de fechamento e vencimento
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS cartoes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome_cartao TEXT,
                nome_usuario TEXT,
                numero TEXT,
                validade TEXT,
                bandeira TEXT,
                limite REAL
            )
        """)
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS metas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT,
                valor_meta REAL,
                valor_atual REAL,
                data_inicial TEXT,
                data_final TEXT
            )
        """)
        self.conn.commit()

    def carregar_dados(self):
        try:
            self.cursor.execute("SELECT ano, mes, despesa, valor, vencimento, categoria, observacao, pagamento, cartao_utilizado FROM despesas")
            self.despesas = self.cursor.fetchall()
            # Incluímos o campo id para os cartões
            self.cursor.execute("SELECT id, nome_cartao, nome_usuario, numero, validade, bandeira, limite FROM cartoes")
            self.cartoes = self.cursor.fetchall()
            self.cursor.execute("SELECT id, nome, valor_meta, valor_atual, data_inicial, data_final FROM metas")
            self.metas = self.cursor.fetchall()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados do banco: {e}")

    # ------------------------------
    # Aba Despesas
    # ------------------------------
    def configurar_aba_despesas(self):
        self.configurar_frame_temas()
        self.configurar_frame_orcamento()
        self.configurar_frame_filtros()
        self.configurar_frame_entrada()
        self.configurar_tabela()
        self.configurar_frame_botoes()

    def configurar_frame_temas(self):
        self.frame_temas = ttk.Frame(self.tab_despesas, padding=10)
        self.frame_temas.pack(fill=ttk.X, pady=5)
        ttk.Label(self.frame_temas, text="Tema:", font=("Helvetica", 10, "bold")).pack(side=ttk.LEFT, padx=5)
        self.tema_var = ttk.StringVar(value="Modo Claro")
        self.combo_tema = ttk.Combobox(self.frame_temas, textvariable=self.tema_var, state="readonly", width=15)
        self.combo_tema["values"] = ["Modo Claro", "Modo Escuro"]
        self.combo_tema.current(0)
        self.combo_tema.pack(side=ttk.LEFT, padx=5)
        ToolTip(self.combo_tema, text="Selecione entre Modo Claro e Modo Escuro.")
        self.combo_tema.bind("<<ComboboxSelected>>", self.mudar_tema)

    def configurar_frame_orcamento(self):
        self.frame_orcamento = ttk.Frame(self.tab_despesas, padding=10)
        self.frame_orcamento.pack(fill=ttk.X, pady=5)
        ttk.Label(self.frame_orcamento, text="Orçamento Mensal:", font=("Helvetica", 10, "bold")).pack(side=ttk.LEFT, padx=5)
        self.orcamento_var = tk.StringVar(value="0,00")
        self.entry_orcamento = ttk.Entry(self.frame_orcamento, textvariable=self.orcamento_var, width=15)
        self.entry_orcamento.pack(side=ttk.LEFT, padx=5)
        ToolTip(self.entry_orcamento, text="Defina o valor do orçamento mensal (use vírgula ou ponto).")
        ttk.Button(self.frame_orcamento, text="Definir Orçamento", bootstyle=SUCCESS, command=self.definir_orcamento).pack(side=ttk.LEFT, padx=5)
        self.progress_bar = ttk.Progressbar(self.frame_orcamento, orient="horizontal", length=200, mode="determinate")
        self.progress_bar.pack(side=ttk.LEFT, padx=10)
        self.label_gastos = ttk.Label(self.frame_orcamento,
                                      text="Gastos: R$ 0,00 / Orçamento: R$ 0,00 / Restante: R$ 0,00",
                                      font=("Helvetica", 10))
        self.label_gastos.pack(side=ttk.LEFT, padx=5)
        ToolTip(self.label_gastos, text="Mostra o total de gastos (somente Débito ou PIX), o orçamento e o saldo restante.")

    def configurar_frame_filtros(self):
        self.frame_filtros = ttk.Frame(self.tab_despesas, padding=10)
        self.frame_filtros.pack(fill=ttk.X, pady=5)
        self.mes_var = ttk.StringVar()
        self.ano_var = ttk.StringVar()
        self.categoria_filtro_var = ttk.StringVar()
        ttk.Label(self.frame_filtros, text="Mês:", font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=5, pady=5, sticky=ttk.E)
        self.combo_mes = ttk.Combobox(self.frame_filtros, textvariable=self.mes_var, state="readonly", width=15)
        self.combo_mes["values"] = [
            "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
            "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
        ]
        self.combo_mes.current(datetime.now().month - 1)
        self.combo_mes.grid(row=0, column=1, padx=5, pady=5, sticky=ttk.W)
        ToolTip(self.combo_mes, text="Selecione o mês para filtrar despesas.")
        ttk.Label(self.frame_filtros, text="Ano:", font=("Helvetica", 10, "bold")).grid(row=0, column=2, padx=5, pady=5, sticky=ttk.E)
        self.combo_ano = ttk.Combobox(self.frame_filtros, textvariable=self.ano_var, state="readonly", width=10)
        self.combo_ano["values"] = [str(ano) for ano in range(2022, datetime.now().year + 2)]
        self.combo_ano.current(datetime.now().year - 2022)
        self.combo_ano.grid(row=0, column=3, padx=5, pady=5, sticky=ttk.W)
        ToolTip(self.combo_ano, text="Selecione o ano para filtrar despesas.")
        ttk.Label(self.frame_filtros, text="Categoria:", font=("Helvetica", 10, "bold")).grid(row=0, column=4, padx=5, pady=5, sticky=ttk.E)
        self.combo_categoria_filtro = ttk.Combobox(self.frame_filtros, textvariable=self.categoria_filtro_var,
                                                   values=["Todas"] + self.categorias, state="readonly", width=15)
        self.combo_categoria_filtro.current(0)
        self.combo_categoria_filtro.grid(row=0, column=5, padx=5, pady=5, sticky=ttk.W)
        ToolTip(self.combo_categoria_filtro, text="Filtre despesas por categoria.")
        self.botao_filtrar = ttk.Button(self.frame_filtros, text="Filtrar", bootstyle=INFO, command=self.filtrar_despesas)
        self.botao_filtrar.grid(row=0, column=6, padx=5, pady=5, sticky=ttk.W)
        ToolTip(self.botao_filtrar, text="Filtre as despesas pelo mês, ano e categoria selecionados.")

    def configurar_frame_entrada(self):
        self.frame_entrada = ttk.Frame(self.tab_despesas, padding=10)
        self.frame_entrada.pack(fill=ttk.X, pady=5)
        self.campos = {}
        labels = ["Despesa", "Valor", "Vencimento", "Observação"]
        placeholders = ["Ex: Supermercado", "Ex: 100,00", "dd/mm/yyyy", "Ex: Compra de Natal"]

        ttk.Label(self.frame_entrada, text=labels[0] + ":", font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=5, pady=5, sticky=ttk.E)
        entry_despesa = ttk.Entry(self.frame_entrada, width=20)
        entry_despesa.grid(row=0, column=1, padx=5, pady=5, sticky=ttk.W)
        entry_despesa.insert(0, placeholders[0])
        entry_despesa.bind("<FocusIn>", lambda event, e=entry_despesa, p=placeholders[0]: self.remover_placeholder(event, e, p))
        entry_despesa.bind("<FocusOut>", lambda event, e=entry_despesa, p=placeholders[0]: self.adicionar_placeholder(event, e, p))
        self.campos[labels[0]] = entry_despesa
        ToolTip(entry_despesa, text="Informe o nome da despesa.")

        ttk.Label(self.frame_entrada, text=labels[1] + ":", font=("Helvetica", 10, "bold")).grid(row=1, column=0, padx=5, pady=5, sticky=ttk.E)
        entry_valor = ttk.Entry(self.frame_entrada, width=20)
        entry_valor.grid(row=1, column=1, padx=5, pady=5, sticky=ttk.W)
        entry_valor.insert(0, placeholders[1])
        entry_valor.bind("<FocusIn>", lambda event, e=entry_valor, p=placeholders[1]: self.remover_placeholder(event, e, p))
        entry_valor.bind("<FocusOut>", lambda event, e=entry_valor, p=placeholders[1]: self.formatar_valor(event, e, p))
        self.campos[labels[1]] = entry_valor
        ToolTip(entry_valor, text="Informe o valor da despesa no formato 100,00 ou 100.00.")

        ttk.Label(self.frame_entrada, text=labels[2] + ":", font=("Helvetica", 10, "bold")).grid(row=2, column=0, padx=5, pady=5, sticky=ttk.E)
        entry_vencimento = ttk.Entry(self.frame_entrada, width=20)
        entry_vencimento.grid(row=2, column=1, padx=5, pady=5, sticky=ttk.W)
        entry_vencimento.insert(0, placeholders[2])
        entry_vencimento.bind("<FocusOut>", lambda event, e=entry_vencimento, p=placeholders[2]: self.formatar_data(event, e, p))
        self.campos[labels[2]] = entry_vencimento
        ToolTip(entry_vencimento, text="Informe a data de vencimento no formato dd/mm/yyyy. (Ex.: 05041999 será convertido para 05/04/1999)")

        ttk.Label(self.frame_entrada, text=labels[3] + ":", font=("Helvetica", 10, "bold")).grid(row=3, column=0, padx=5, pady=5, sticky=ttk.E)
        entry_observacao = ttk.Entry(self.frame_entrada, width=20)
        entry_observacao.grid(row=3, column=1, padx=5, pady=5, sticky=ttk.W)
        entry_observacao.insert(0, placeholders[3])
        entry_observacao.bind("<FocusIn>", lambda event, e=entry_observacao, p=placeholders[3]: self.remover_placeholder(event, e, p))
        entry_observacao.bind("<FocusOut>", lambda event, e=entry_observacao, p=placeholders[3]: self.adicionar_placeholder(event, e, p))
        self.campos[labels[3]] = entry_observacao
        ToolTip(entry_observacao, text="Adicione uma observação ou descrição para a despesa.")

        ttk.Label(self.frame_entrada, text="Categoria:", font=("Helvetica", 10, "bold")).grid(row=4, column=0, padx=5, pady=5, sticky=ttk.E)
        self.categoria_var = ttk.StringVar()
        self.combo_categoria = ttk.Combobox(self.frame_entrada, textvariable=self.categoria_var,
                                            values=self.categorias, state="readonly", width=18)
        self.combo_categoria.grid(row=4, column=1, padx=5, pady=5, sticky=ttk.W)
        self.combo_categoria.current(0)
        ToolTip(self.combo_categoria, text="Selecione a categoria da despesa.")

        ttk.Label(self.frame_entrada, text="Forma de Pagamento:", font=("Helvetica", 10, "bold")).grid(row=5, column=0, padx=5, pady=5, sticky=ttk.E)
        self.forma_pagamento_var = ttk.StringVar()
        self.combo_pagamento = ttk.Combobox(self.frame_entrada, textvariable=self.forma_pagamento_var,
                                            values=self.formas_pagamento, state="readonly", width=18)
        self.combo_pagamento.grid(row=5, column=1, padx=5, pady=5, sticky=ttk.W)
        self.combo_pagamento.current(0)
        ToolTip(self.combo_pagamento, text="Selecione a forma de pagamento (VR, Cartão de Crédito, PIX ou Débito).")

        ttk.Label(self.frame_entrada, text="Cartão Utilizado:", font=("Helvetica", 10, "bold")).grid(row=6, column=0, padx=5, pady=5, sticky=ttk.E)
        self.cartao_utilizado_var = ttk.StringVar()
        self.combo_cartao_utilizado = ttk.Combobox(self.frame_entrada, textvariable=self.cartao_utilizado_var,
                                                   values=[c[1] for c in self.cartoes] if self.cartoes else [],
                                                   state="readonly", width=18)
        self.combo_cartao_utilizado.grid(row=6, column=1, padx=5, pady=5, sticky=ttk.W)
        ToolTip(self.combo_cartao_utilizado, text="Selecione o cartão utilizado (se for Cartão de Crédito).")

        self.botao_adicionar = ttk.Button(self.frame_entrada, text="Adicionar Despesa", bootstyle=SUCCESS, command=self.adicionar_despesa)
        self.botao_adicionar.grid(row=7, column=0, columnspan=2, pady=10)
        ToolTip(self.botao_adicionar, text="Adicione uma nova despesa à lista.")

    def configurar_tabela(self):
        self.frame_tabela = ttk.Frame(self.tab_despesas)
        self.frame_tabela.pack(fill=ttk.BOTH, expand=True, pady=10)
        colunas = ("Ano", "Mês", "Despesa", "Valor", "Vencimento", "Categoria", "Observação", "Pagamento", "Cartão Utilizado")
        self.tabela = ttk.Treeview(self.frame_tabela, columns=colunas, show="headings")
        self.tabela.pack(fill=ttk.BOTH, expand=True)
        for col in colunas:
            self.tabela.heading(col, text=col)
            self.tabela.column(col, width=120, anchor="center")

    def configurar_frame_botoes(self):
        self.frame_botoes = ttk.Frame(self.tab_despesas, padding=10)
        self.frame_botoes.pack(fill=ttk.X)
        self.botao_excluir = ttk.Button(self.frame_botoes, text="Excluir Despesa", bootstyle=DANGER, command=self.excluir_despesa)
        self.botao_excluir.pack(side=ttk.LEFT, padx=5)
        ToolTip(self.botao_excluir, text="Exclua a despesa selecionada.")
        self.botao_salvar_excel = ttk.Button(self.frame_botoes, text="Salvar no Excel", bootstyle=PRIMARY, command=self.salvar_no_excel)
        self.botao_salvar_excel.pack(side=ttk.LEFT, padx=5)
        ToolTip(self.botao_salvar_excel, text="Exporte os dados para um arquivo Excel.")
        self.botao_importar = ttk.Button(self.frame_botoes, text="Importar Planilha", bootstyle=INFO, command=self.importar_planilha)
        self.botao_importar.pack(side=ttk.LEFT, padx=5)
        ToolTip(self.botao_importar, text="Importe despesas de uma planilha Excel.")
        self.botao_backup = ttk.Button(self.frame_botoes, text="Backup DB", bootstyle=INFO, command=self.backup_db)
        self.botao_backup.pack(side=ttk.LEFT, padx=5)
        ToolTip(self.botao_backup, text="Realize um backup do banco de dados.")
        self.botao_restaurar = ttk.Button(self.frame_botoes, text="Restaurar Backup", bootstyle=WARNING, command=self.restaurar_backup)
        self.botao_restaurar.pack(side=ttk.LEFT, padx=5)
        ToolTip(self.botao_restaurar, text="Restaure um backup do banco de dados.")
        self.botao_sair = ttk.Button(self.frame_botoes, text="Sair", bootstyle=SECONDARY, command=self.destroy)
        self.botao_sair.pack(side=ttk.RIGHT, padx=5)
        ToolTip(self.botao_sair, text="Feche o aplicativo.")

    # -------------------------
    # Lógica de Despesas
    # -------------------------
    def mudar_tema(self, event=None):
        if self.tema_var.get() == "Modo Claro":
            novo_tema = "cosmo"
        else:
            novo_tema = "darkly"
        self.style.theme_use(novo_tema)

    def definir_orcamento(self):
        try:
            valor_texto = self.orcamento_var.get().replace(',', '.')
            self.orcamento_mensal = float(valor_texto)
            messagebox.showinfo("Sucesso", f"Orçamento mensal definido como R$ {self.orcamento_var.get()}")
            self.atualizar_indicador_gastos()
        except ValueError:
            messagebox.showerror("Erro", "Insira um valor numérico válido para o orçamento.")

    def atualizar_indicador_gastos(self):
        total_gastos = sum(float(d[3]) for d in self.despesas if d[7] in ["Débito", "PIX"])
        restante = self.orcamento_mensal - total_gastos
        if restante < 0:
            restante = 0
        txt_gastos = f"{total_gastos:.2f}".replace('.', ',')
        txt_orc = self.orcamento_var.get()
        txt_rest = f"{restante:.2f}".replace('.', ',')
        if self.orcamento_mensal > 0:
            percentual_gasto = (total_gastos / self.orcamento_mensal) * 100
            self.progress_bar["value"] = percentual_gasto
            self.label_gastos.config(
                text=f"Gastos: R$ {txt_gastos} / Orçamento: R$ {txt_orc} / Restante: R$ {txt_rest}"
            )
        else:
            self.progress_bar["value"] = 0
            self.label_gastos.config(
                text=f"Gastos: R$ {txt_gastos} / Orçamento: R$ 0,00 / Restante: R$ 0,00"
            )

    def adicionar_despesa(self):
        valores = [campo.get() for campo in self.campos.values()]
        categoria = self.categoria_var.get()
        forma_pagamento = self.forma_pagamento_var.get()
        cartao_utilizado = self.cartao_utilizado_var.get() if forma_pagamento == "Cartão de Crédito" else ""
        if not all(valores[:3]):
            messagebox.showerror("Erro", "Preencha os campos obrigatórios (Despesa, Valor, Vencimento).")
            return
        try:
            valor = float(valores[1].replace(",", "."))
        except ValueError:
            messagebox.showerror("Erro", "O campo Valor deve ser numérico.")
            return
        try:
            datetime.strptime(valores[2], "%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Erro", "Data de vencimento inválida. Use o formato dd/mm/yyyy.")
            return
        nova_despesa = (
            self.combo_ano.get(),
            self.combo_mes.get(),
            valores[0],
            valor,
            valores[2],
            categoria,
            valores[3],
            forma_pagamento,
            cartao_utilizado
        )
        self.despesas.append(nova_despesa)
        try:
            self.cursor.execute(
                "INSERT INTO despesas (ano, mes, despesa, valor, vencimento, categoria, observacao, pagamento, cartao_utilizado) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                nova_despesa
            )
            self.conn.commit()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao inserir despesa no banco: {e}")
        self.atualizar_tabela()
        self.atualizar_indicador_gastos()
        self.atualizar_tabela_despesas_cartao()
        for campo in self.campos.values():
            campo.delete(0, "end")
        self.categoria_var.set("")
        self.forma_pagamento_var.set(self.formas_pagamento[0])
        self.combo_cartao_utilizado["values"] = [c[1] for c in self.cartoes]

    def excluir_despesa(self):
        selecionado = self.tabela.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione uma despesa para excluir.")
            return
        for item in selecionado:
            valores = self.tabela.item(item, "values")
            self.despesas = [d for d in self.despesas if d != valores]
            self.tabela.delete(item)
            try:
                self.cursor.execute("""
                    DELETE FROM despesas 
                    WHERE ano=? AND mes=? AND despesa=? AND valor=? AND vencimento=? AND categoria=? AND observacao=? AND pagamento=? AND cartao_utilizado=?
                """, valores)
                self.conn.commit()
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao excluir despesa do banco: {e}")
        self.atualizar_indicador_gastos()
        self.atualizar_tabela_despesas_cartao()

    def filtrar_despesas(self):
        mes_filtro = self.mes_var.get()
        ano_filtro = self.ano_var.get()
        categoria_filtro = self.categoria_filtro_var.get()
        self.tabela.delete(*self.tabela.get_children())
        for despesa in self.despesas:
            if (despesa[0] == ano_filtro or ano_filtro == "") and \
               (despesa[1] == mes_filtro or mes_filtro == "") and \
               (despesa[5] == categoria_filtro or categoria_filtro == "Todas"):
                self.tabela.insert("", "end", values=despesa)

    def salvar_no_excel(self):
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                filetypes=[("Excel Files", "*.xlsx")])
        if not filepath:
            return
        workbook = openpyxl.Workbook()
        # Aba Despesas
        despesas_sheet = workbook.active
        despesas_sheet.title = "Despesas"
        cabecalho = ["Ano", "Mês", "Despesa", "Valor", "Vencimento", "Categoria", "Observação", "Pagamento", "Cartão Utilizado"]
        despesas_sheet.append(cabecalho)
        for row in self.despesas:
            despesas_sheet.append(row)
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center")
        header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        for col in range(1, despesas_sheet.max_column + 1):
            cell = despesas_sheet.cell(row=1, column=col)
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.fill = header_fill
            cell.border = thin_border
        despesas_sheet.auto_filter.ref = f"A1:I1"
        for row in despesas_sheet.iter_rows(min_row=2, max_row=despesas_sheet.max_row, min_col=1, max_col=despesas_sheet.max_column):
            for cell in row:
                cell.border = thin_border
        total_gastos = sum(float(d[3]) for d in self.despesas if d[7] in ["Débito", "PIX"])
        saldo_restante = self.orcamento_mensal - total_gastos
        if saldo_restante < 0:
            saldo_restante = 0
        despesas_sheet.append([])
        despesas_sheet.append(["Orçamento", f"R$ {self.orcamento_var.get()}"])
        despesas_sheet.append(["Saldo Restante", f"R$ {saldo_restante:.2f}".replace('.', ',')])
        for col in despesas_sheet.columns:
            max_length = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            despesas_sheet.column_dimensions[col_letter].width = max_length + 2

        # Aba Cartões
        cartoes_sheet = workbook.create_sheet(title="Cartões")
        cabecalho_cartoes = ["Nome do Cartão", "Nome do Usuário", "Número", "Validade", "Bandeira", "Limite"]
        cartoes_sheet.append(cabecalho_cartoes)
        for cartao in self.cartoes:
            # Exibe os dados sem o id
            cartao_exibicao = (cartao[1], cartao[2], cartao[3], cartao[4], cartao[5], cartao[6])
            cartoes_sheet.append(cartao_exibicao)
        for col in range(1, cartoes_sheet.max_column + 1):
            cell = cartoes_sheet.cell(row=1, column=col)
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.fill = header_fill
            cell.border = thin_border
        cartoes_sheet.auto_filter.ref = f"A1:F1"
        for row in cartoes_sheet.iter_rows(min_row=2, max_row=cartoes_sheet.max_row, min_col=1, max_col=cartoes_sheet.max_column):
            for cell in row:
                cell.border = thin_border

        # Bloco: Despesas no Cartão de Crédito
        cartoes_sheet.append([])
        cartoes_sheet.append(["Despesas no Cartão de Crédito"])
        cabecalho_desp_cartao = ["Ano", "Mês", "Despesa", "Valor", "Vencimento", "Categoria", "Observação", "Pagamento", "Cartão Utilizado"]
        cartoes_sheet.append(cabecalho_desp_cartao)
        for despesa in self.despesas:
            if despesa[7] == "Cartão de Crédito":
                cartoes_sheet.append(despesa)
        for col in range(1, len(cabecalho_desp_cartao) + 1):
            linha_cab = cartoes_sheet.max_row - sum(1 for d in self.despesas if d[7]=="Cartão de Crédito") - 1
            cell = cartoes_sheet.cell(row=linha_cab, column=col)
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.fill = header_fill
            cell.border = thin_border
        for col in cartoes_sheet.columns:
            max_length = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            cartoes_sheet.column_dimensions[col_letter].width = max_length + 2

        workbook.save(filepath)
        messagebox.showinfo("Sucesso", f"Dados salvos com sucesso em {filepath}!")

    def atualizar_tabela(self):
        self.tabela.delete(*self.tabela.get_children())
        for despesa in self.despesas:
            self.tabela.insert("", "end", values=despesa)

    def atualizar_tabela_despesas_cartao(self):
        self.tabela_despesas_cartao.delete(*self.tabela_despesas_cartao.get_children())
        for despesa in self.despesas:
            if despesa[7] == "Cartão de Crédito":
                self.tabela_despesas_cartao.insert("", "end", values=despesa)

    # -------------------------
    # Aba Cartão de Crédito
    # -------------------------
    def configurar_aba_cartao(self):
        self.frame_cartao_form = ttk.Frame(self.tab_cartao, padding=10)
        self.frame_cartao_form.pack(fill=ttk.X, pady=5)
        ttk.Label(self.frame_cartao_form, text="Nome do Cartão:", font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=5, pady=5, sticky=tk.E)
        self.entry_nome_cartao = ttk.Entry(self.frame_cartao_form, width=25)
        self.entry_nome_cartao.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        ToolTip(self.entry_nome_cartao, text="Ex: 'Meu Nubank', 'Cartão Santander' etc.")
        ttk.Label(self.frame_cartao_form, text="Nome do Usuário:", font=("Helvetica", 10, "bold")).grid(row=1, column=0, padx=5, pady=5, sticky=tk.E)
        self.entry_nome_usuario = ttk.Entry(self.frame_cartao_form, width=25)
        self.entry_nome_usuario.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)
        ToolTip(self.entry_nome_usuario, text="Digite o nome do usuário do cartão.")
        ttk.Label(self.frame_cartao_form, text="Número do Cartão:", font=("Helvetica", 10, "bold")).grid(row=2, column=0, padx=5, pady=5, sticky=tk.E)
        self.entry_numero_cartao = ttk.Entry(self.frame_cartao_form, width=25)
        self.entry_numero_cartao.grid(row=2, column=1, padx=5, pady=5, sticky=tk.W)
        ToolTip(self.entry_numero_cartao, text="Digite o número do cartão (sem mascaramento).")
        ttk.Label(self.frame_cartao_form, text="Data de Validade (mm/aa):", font=("Helvetica", 10, "bold")).grid(row=3, column=0, padx=5, pady=5, sticky=tk.E)
        self.entry_validade = ttk.Entry(self.frame_cartao_form, width=10)
        self.entry_validade.grid(row=3, column=1, padx=5, pady=5, sticky=tk.W)
        ToolTip(self.entry_validade, text="Digite a data de validade no formato mm/aa.")
        ttk.Label(self.frame_cartao_form, text="Bandeira:", font=("Helvetica", 10, "bold")).grid(row=4, column=0, padx=5, pady=5, sticky=tk.E)
        self.combo_bandeira = ttk.Combobox(self.frame_cartao_form, state="readonly", width=15)
        self.combo_bandeira["values"] = ["Visa", "MasterCard", "American Express", "Elo", "Hipercard", "Outros"]
        self.combo_bandeira.current(0)
        self.combo_bandeira.grid(row=4, column=1, padx=5, pady=5, sticky=tk.W)
        ToolTip(self.combo_bandeira, text="Selecione a bandeira do cartão.")
        ttk.Label(self.frame_cartao_form, text="Limite de Crédito (R$):", font=("Helvetica", 10, "bold")).grid(row=5, column=0, padx=5, pady=5, sticky=tk.E)
        self.entry_limite = ttk.Entry(self.frame_cartao_form, width=15)
        self.entry_limite.grid(row=5, column=1, padx=5, pady=5, sticky=tk.W)
        ToolTip(self.entry_limite, text="Digite o limite de crédito do cartão (ex: 1000,00).")
        
        # Botões de ação para Cartão
        self.frame_cartao_botoes = ttk.Frame(self.tab_cartao, padding=10)
        self.frame_cartao_botoes.pack(fill=ttk.X, pady=5)
        self.botao_cadastrar_cartao = ttk.Button(self.frame_cartao_botoes, text="Cadastrar Cartão", bootstyle=SUCCESS, command=self.cadastrar_cartao)
        self.botao_cadastrar_cartao.pack(side=ttk.LEFT, padx=5)
        self.botao_editar_cartao = ttk.Button(self.frame_cartao_botoes, text="Editar Cartão", bootstyle=INFO, command=self.atualizar_cartao)
        self.botao_editar_cartao.pack(side=ttk.LEFT, padx=5)
        self.botao_excluir_cartao = ttk.Button(self.frame_cartao_botoes, text="Excluir Cartão", bootstyle=DANGER, command=self.excluir_cartao)
        self.botao_excluir_cartao.pack(side=ttk.LEFT, padx=5)

        self.frame_cartao_tabela = ttk.Frame(self.tab_cartao, padding=10)
        self.frame_cartao_tabela.pack(fill=ttk.BOTH, expand=True, pady=5)
        colunas = ("Nome do Cartão", "Nome do Usuário", "Número", "Validade", "Bandeira", "Limite")
        self.tabela_cartao = ttk.Treeview(self.frame_cartao_tabela, columns=colunas, show="headings")
        self.tabela_cartao.pack(fill=ttk.BOTH, expand=True)
        for col in colunas:
            self.tabela_cartao.heading(col, text=col)
            self.tabela_cartao.column(col, width=120, anchor="center")
        # Vincula seleção para carregar dados no formulário
        self.tabela_cartao.bind("<<TreeviewSelect>>", self.carregar_cartao_selecionado)

        self.frame_dashboard = ttk.Frame(self.tab_cartao, padding=10)
        self.frame_dashboard.pack(fill=ttk.X, pady=5)
        ttk.Label(self.frame_dashboard, text="Dashboard do Cartão:", font=("Helvetica", 12, "bold")).pack(side=ttk.LEFT, padx=5)
        self.cartao_dashboard_var = ttk.StringVar()
        self.combo_dashboard = ttk.Combobox(self.frame_dashboard, textvariable=self.cartao_dashboard_var, state="readonly", width=25)
        self.combo_dashboard["values"] = [c[1] for c in self.cartoes] if self.cartoes else []
        if self.cartoes:
            self.combo_dashboard.current(0)
        self.combo_dashboard.pack(side=ttk.LEFT, padx=5)
        self.combo_dashboard.bind("<<ComboboxSelected>>", self.atualizar_dashboard_cartao)
        self.progress_cartao = ttk.Progressbar(self.frame_dashboard, orient="horizontal", length=200, mode="determinate")
        self.progress_cartao.pack(side=ttk.LEFT, padx=5)
        self.label_dashboard = ttk.Label(self.frame_dashboard, text="Gastos: R$ 0,00 / Limite: R$ 0,00 / Disponível: R$ 0,00", font=("Helvetica", 10))
        self.label_dashboard.pack(side=ttk.LEFT, padx=5)

        ttk.Label(self.tab_cartao, text="Despesas no Cartão de Crédito", font=("Helvetica", 12, "bold")).pack(pady=10)
        colunas = ("Ano", "Mês", "Despesa", "Valor", "Vencimento", "Categoria", "Observação", "Pagamento", "Cartão Utilizado")
        self.tabela_despesas_cartao = ttk.Treeview(self.tab_cartao, columns=colunas, show="headings")
        self.tabela_despesas_cartao.pack(fill=ttk.BOTH, expand=True, pady=5)
        for col in colunas:
            self.tabela_despesas_cartao.heading(col, text=col)
            self.tabela_despesas_cartao.column(col, width=120, anchor="center")

    def cadastrar_cartao(self):
        nome_cartao = self.entry_nome_cartao.get().strip()
        nome_usuario = self.entry_nome_usuario.get().strip()
        numero = self.entry_numero_cartao.get().strip()
        validade = self.entry_validade.get().strip()
        bandeira = self.combo_bandeira.get().strip()
        limite_texto = self.entry_limite.get().strip().replace(',', '.')
        if not nome_cartao or not nome_usuario or not numero or not validade or not bandeira or not limite_texto:
            messagebox.showerror("Erro", "Preencha todos os campos obrigatórios.")
            return
        try:
            datetime.strptime(validade, "%m/%y")
        except ValueError:
            messagebox.showerror("Erro", "Data de validade inválida. Use o formato mm/aa.")
            return
        try:
            limite = float(limite_texto)
        except ValueError:
            messagebox.showerror("Erro", "O campo Limite de Crédito deve ser numérico.")
            return
        if len(numero) >= 4:
            numero_mascarado = "**** " + numero[-4:]
        else:
            numero_mascarado = numero
        # Insere os dados do cartão (sem fechamento e vencimento)
        cartao = (
            nome_cartao,
            nome_usuario,
            numero_mascarado,
            validade,
            bandeira,
            f"{limite:.2f}".replace('.', ',')
        )
        try:
            self.cursor.execute(
                "INSERT INTO cartoes (nome_cartao, nome_usuario, numero, validade, bandeira, limite) VALUES (?, ?, ?, ?, ?, ?)",
                cartao
            )
            self.conn.commit()
            # Recupera o id recém-inserido e atualiza a lista de cartões
            card_id = self.cursor.lastrowid
            self.cartoes.append((card_id,) + cartao)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao inserir cartão no banco: {e}")
            return
        self.atualizar_tabela_cartao()
        messagebox.showinfo("Sucesso", "Cartão cadastrado com sucesso!")
        self.entry_nome_cartao.delete(0, "end")
        self.entry_nome_usuario.delete(0, "end")
        self.entry_numero_cartao.delete(0, "end")
        self.entry_validade.delete(0, "end")
        self.combo_bandeira.current(0)
        self.entry_limite.delete(0, "end")
        self.combo_cartao_utilizado["values"] = [c[1] for c in self.cartoes]
        self.combo_dashboard["values"] = [c[1] for c in self.cartoes]
        if self.cartoes:
            self.combo_dashboard.current(0)
            self.atualizar_dashboard_cartao()

    def carregar_cartao_selecionado(self, event):
        selected = self.tabela_cartao.selection()
        if not selected:
            return
        card_id = int(selected[0])
        # Procura o cartão com esse id
        for card in self.cartoes:
            if card[0] == card_id:
                self.card_edit_id = card_id
                self.entry_nome_cartao.delete(0, tk.END)
                self.entry_nome_cartao.insert(0, card[1])
                self.entry_nome_usuario.delete(0, tk.END)
                self.entry_nome_usuario.insert(0, card[2])
                # Note que o número já está mascarado
                self.entry_numero_cartao.delete(0, tk.END)
                self.entry_numero_cartao.insert(0, card[3])
                self.entry_validade.delete(0, tk.END)
                self.entry_validade.insert(0, card[4])
                self.combo_bandeira.set(card[5])
                self.entry_limite.delete(0, tk.END)
                self.entry_limite.insert(0, card[6])
                break

    def atualizar_cartao(self):
        # Verifica se um cartão foi selecionado para edição
        if not hasattr(self, "card_edit_id"):
            messagebox.showwarning("Aviso", "Selecione um cartão para editar.")
            return
        card_id = self.card_edit_id
        nome_cartao = self.entry_nome_cartao.get().strip()
        nome_usuario = self.entry_nome_usuario.get().strip()
        numero = self.entry_numero_cartao.get().strip()
        validade = self.entry_validade.get().strip()
        bandeira = self.combo_bandeira.get().strip()
        limite_texto = self.entry_limite.get().strip().replace(',', '.')
        if not nome_cartao or not nome_usuario or not numero or not validade or not bandeira or not limite_texto:
            messagebox.showerror("Erro", "Preencha todos os campos obrigatórios para edição.")
            return
        try:
            datetime.strptime(validade, "%m/%y")
        except ValueError:
            messagebox.showerror("Erro", "Data de validade inválida. Use o formato mm/aa.")
            return
        try:
            limite = float(limite_texto)
        except ValueError:
            messagebox.showerror("Erro", "O campo Limite de Crédito deve ser numérico.")
            return
        if len(numero) >= 4:
            numero_mascarado = "**** " + numero[-4:]
        else:
            numero_mascarado = numero
        try:
            self.cursor.execute(
                "UPDATE cartoes SET nome_cartao=?, nome_usuario=?, numero=?, validade=?, bandeira=?, limite=? WHERE id=?",
                (nome_cartao, nome_usuario, numero_mascarado, validade, bandeira, f"{limite:.2f}".replace('.', ','), card_id)
            )
            self.conn.commit()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao atualizar cartão: {e}")
            return
        # Atualiza a lista de cartões
        for i, card in enumerate(self.cartoes):
            if card[0] == card_id:
                self.cartoes[i] = (card_id, nome_cartao, nome_usuario, numero_mascarado, validade, bandeira, f"{limite:.2f}".replace('.', ','))
                break
        self.atualizar_tabela_cartao()
        messagebox.showinfo("Sucesso", "Cartão atualizado com sucesso!")
        # Limpa a variável de edição
        self.card_edit_id = None

    def excluir_cartao(self):
        selected = self.tabela_cartao.selection()
        if not selected:
            messagebox.showwarning("Aviso", "Selecione um cartão para excluir.")
            return
        card_id = int(selected[0])
        try:
            self.cursor.execute("DELETE FROM cartoes WHERE id=?", (card_id,))
            self.conn.commit()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao excluir cartão: {e}")
            return
        # Remove da lista
        self.cartoes = [c for c in self.cartoes if c[0] != card_id]
        self.atualizar_tabela_cartao()
        messagebox.showinfo("Sucesso", "Cartão excluído com sucesso!")
        # Atualiza os comboboxes que usam a lista de cartões
        self.combo_cartao_utilizado["values"] = [c[1] for c in self.cartoes]
        self.combo_dashboard["values"] = [c[1] for c in self.cartoes]

    def atualizar_tabela_cartao(self):
        self.tabela_cartao.delete(*self.tabela_cartao.get_children())
        for card in self.cartoes:
            # Exibe sem o id; usa o id como iid
            self.tabela_cartao.insert("", "end", iid=str(card[0]), values=(card[1], card[2], card[3], card[4], card[5], card[6]))

    def atualizar_dashboard_cartao(self, event=None):
        selected_card = self.cartao_dashboard_var.get()
        card = None
        for c in self.cartoes:
            if c[1].strip() == selected_card.strip():
                card = c
                break
        if not card:
            return
        # Aqui, como não temos fechamento e vencimento, o dashboard pode mostrar apenas o total gasto (calculado a partir das despesas com o nome do cartão)
        limite = float(card[6].replace(',', '.'))
        total_cartao = sum(float(d[3]) for d in self.despesas if d[7]=="Cartão de Crédito" and d[8].strip() == card[1].strip())
        disponivel = limite - total_cartao
        if disponivel < 0:
            disponivel = 0
        percentual = (total_cartao / limite * 100) if limite > 0 else 0
        self.progress_cartao["value"] = percentual
        self.label_dashboard.config(
            text=(f"Gastos: R$ {total_cartao:.2f}".replace('.', ',') +
                  f" / Limite: R$ {card[6]} / Disponível: R$ {disponivel:.2f}".replace('.', ','))
        )
        if percentual >= 90:
            messagebox.showwarning("Alerta", f"Você atingiu {percentual:.0f}% do limite do cartão {selected_card}!")

    def carregar_cartao_selecionado(self, event):
        selected = self.tabela_cartao.selection()
        if not selected:
            return
        card_id = int(selected[0])
        for card in self.cartoes:
            if card[0] == card_id:
                self.card_edit_id = card_id
                self.entry_nome_cartao.delete(0, tk.END)
                self.entry_nome_cartao.insert(0, card[1])
                self.entry_nome_usuario.delete(0, tk.END)
                self.entry_nome_usuario.insert(0, card[2])
                self.entry_numero_cartao.delete(0, tk.END)
                self.entry_numero_cartao.insert(0, card[3])
                self.entry_validade.delete(0, tk.END)
                self.entry_validade.insert(0, card[4])
                self.combo_bandeira.set(card[5])
                self.entry_limite.delete(0, tk.END)
                self.entry_limite.insert(0, card[6])
                break

    # -------------------------
    # Funções Genéricas (Data, Placeholders, Valor)
    # -------------------------
    def formatar_data(self, event, entry, placeholder):
        text = entry.get().strip()
        if not text:
            entry.insert(0, placeholder)
            return
        if "/" not in text and len(text) == 8 and text.isdigit():
            try:
                formatted = f"{text[:2]}/{text[2:4]}/{text[4:]}"
                datetime.strptime(formatted, "%d/%m/%Y")
                entry.delete(0, "end")
                entry.insert(0, formatted)
            except Exception:
                pass
        elif "/" in text:
            try:
                parts = text.split("/")
                if len(parts) == 3:
                    day = parts[0].zfill(2)
                    month = parts[1].zfill(2)
                    year = parts[2]
                    if len(year) == 2:
                        year = "20" + year
                    formatted = f"{day}/{month}/{year}"
                    datetime.strptime(formatted, "%d/%m/%Y")
                    entry.delete(0, "end")
                    entry.insert(0, formatted)
            except Exception:
                pass

    def remover_placeholder(self, event, entry, placeholder):
        if entry.get() == placeholder:
            entry.delete(0, "end")

    def adicionar_placeholder(self, event, entry, placeholder):
        if not entry.get():
            entry.insert(0, placeholder)

    def formatar_valor(self, event, entry, placeholder):
        if entry.get() == placeholder:
            return
        try:
            valor = float(entry.get().replace(",", "."))
            entry.delete(0, "end")
            entry.insert(0, f"{valor:.2f}".replace(".", ","))
        except ValueError:
            messagebox.showerror("Erro", "Insira um valor numérico válido.")
            entry.delete(0, "end")
            entry.insert(0, placeholder)

    # -------------------------
    # Aba Relatórios
    # -------------------------
    def configurar_aba_relatorios(self):
        self.frame_relatorios = ttk.Frame(self.tab_relatorios, padding=10)
        self.frame_relatorios.pack(fill=ttk.BOTH, expand=True)
        self.botao_gerar_relatorio = ttk.Button(self.frame_relatorios, text="Gerar Relatório", bootstyle=INFO, command=self.gerar_relatorio)
        self.botao_gerar_relatorio.pack(pady=10)
        self.text_relatorio = tk.Text(self.frame_relatorios, wrap="word", height=20)
        self.text_relatorio.pack(fill=ttk.BOTH, expand=True)

    def gerar_relatorio(self):
        relatorio = "Relatório de Despesas\n\n"
        total_geral = sum(float(d[3]) for d in self.despesas)
        relatorio += f"Total Geral: R$ {total_geral:.2f}\n\n"
        por_categoria = {}
        for d in self.despesas:
            cat = d[5]
            valor = float(d[3])
            por_categoria[cat] = por_categoria.get(cat, 0) + valor
        for cat, total in por_categoria.items():
            relatorio += f"{cat}: R$ {total:.2f}\n"
        self.text_relatorio.delete(1.0, tk.END)
        self.text_relatorio.insert(tk.END, relatorio)

    # -------------------------
    # Aba Metas
    # -------------------------
    def configurar_aba_metas(self):
        self.frame_metas_form = ttk.Frame(self.tab_metas, padding=10)
        self.frame_metas_form.pack(fill=ttk.X, pady=5)
        ttk.Label(self.frame_metas_form, text="Nome da Meta:", font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=5, pady=5, sticky=tk.E)
        self.entry_meta_nome = ttk.Entry(self.frame_metas_form, width=25)
        self.entry_meta_nome.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        ttk.Label(self.frame_metas_form, text="Valor Meta (R$):", font=("Helvetica", 10, "bold")).grid(row=1, column=0, padx=5, pady=5, sticky=tk.E)
        self.entry_meta_valor = ttk.Entry(self.frame_metas_form, width=15)
        self.entry_meta_valor.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)
        ttk.Label(self.frame_metas_form, text="Valor Atual (R$):", font=("Helvetica", 10, "bold")).grid(row=2, column=0, padx=5, pady=5, sticky=tk.E)
        self.entry_meta_atual = ttk.Entry(self.frame_metas_form, width=15)
        self.entry_meta_atual.grid(row=2, column=1, padx=5, pady=5, sticky=tk.W)
        ttk.Label(self.frame_metas_form, text="Data Inicial:", font=("Helvetica", 10, "bold")).grid(row=3, column=0, padx=5, pady=5, sticky=tk.E)
        self.entry_meta_data_inicial = ttk.Entry(self.frame_metas_form, width=15)
        self.entry_meta_data_inicial.grid(row=3, column=1, padx=5, pady=5, sticky=tk.W)
        self.entry_meta_data_inicial.insert(0, "dd/mm/aaaa")
        ttk.Label(self.frame_metas_form, text="Data Final:", font=("Helvetica", 10, "bold")).grid(row=4, column=0, padx=5, pady=5, sticky=tk.E)
        self.entry_meta_data_final = ttk.Entry(self.frame_metas_form, width=15)
        self.entry_meta_data_final.grid(row=4, column=1, padx=5, pady=5, sticky=tk.W)
        self.entry_meta_data_final.insert(0, "dd/mm/aaaa")
        self.botao_cadastrar_meta = ttk.Button(self.frame_metas_form, text="Cadastrar Meta", bootstyle=SUCCESS, command=self.cadastrar_meta)
        self.botao_cadastrar_meta.grid(row=5, column=0, padx=5, pady=10)
        self.botao_atualizar_meta = ttk.Button(self.frame_metas_form, text="Atualizar Meta", bootstyle=INFO, command=self.atualizar_meta)
        self.botao_atualizar_meta.grid(row=5, column=1, padx=5, pady=10)
        self.frame_metas_tabela = ttk.Frame(self.tab_metas, padding=10)
        self.frame_metas_tabela.pack(fill=ttk.BOTH, expand=True, pady=5)
        colunas = ("Nome", "Valor Meta", "Valor Atual", "Progresso (%)", "Data Inicial", "Data Final")
        self.tabela_metas = ttk.Treeview(self.frame_metas_tabela, columns=colunas, show="headings")
        self.tabela_metas.pack(fill=ttk.BOTH, expand=True)
        for col in colunas:
            self.tabela_metas.heading(col, text=col)
            self.tabela_metas.column(col, width=120, anchor="center")
        self.tabela_metas.bind("<<TreeviewSelect>>", self.carregar_meta_selecionada)
        self.atualizar_tabela_metas()

    def cadastrar_meta(self):
        nome = self.entry_meta_nome.get().strip()
        valor_meta_texto = self.entry_meta_valor.get().strip().replace(",", ".")
        valor_atual_texto = self.entry_meta_atual.get().strip().replace(",", ".")
        data_inicial = self.entry_meta_data_inicial.get().strip()
        data_final = self.entry_meta_data_final.get().strip()
        if not nome or not valor_meta_texto or not data_inicial or not data_final:
            messagebox.showerror("Erro", "Preencha os campos obrigatórios da meta (Nome, Valor Meta, Data Inicial e Data Final).")
            return
        try:
            valor_meta = float(valor_meta_texto)
        except ValueError:
            messagebox.showerror("Erro", "O campo Valor Meta deve ser numérico.")
            return
        try:
            valor_atual = float(valor_atual_texto) if valor_atual_texto else 0.0
        except ValueError:
            messagebox.showerror("Erro", "O campo Valor Atual deve ser numérico.")
            return
        try:
            datetime.strptime(data_inicial, "%d/%m/%Y")
            datetime.strptime(data_final, "%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Erro", "Data Inicial ou Data Final inválida. Use o formato dd/mm/aaaa.")
            return
        try:
            self.cursor.execute(
                "INSERT INTO metas (nome, valor_meta, valor_atual, data_inicial, data_final) VALUES (?, ?, ?, ?, ?)",
                (nome, valor_meta, valor_atual, data_inicial, data_final)
            )
            self.conn.commit()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao inserir meta no banco: {e}")
            return
        self.cursor.execute("SELECT id, nome, valor_meta, valor_atual, data_inicial, data_final FROM metas")
        self.metas = self.cursor.fetchall()
        self.atualizar_tabela_metas()
        messagebox.showinfo("Sucesso", "Meta cadastrada com sucesso!")
        self.limpar_form_meta()

    def carregar_meta_selecionada(self, event):
        selected = self.tabela_metas.selection()
        if not selected:
            return
        meta_id = self.tabela_metas.item(selected[0])['iid']
        valores = self.tabela_metas.item(selected[0])['values']
        self.entry_meta_nome.delete(0, tk.END)
        self.entry_meta_nome.insert(0, valores[0])
        self.entry_meta_valor.delete(0, tk.END)
        self.entry_meta_valor.insert(0, valores[1])
        self.entry_meta_atual.delete(0, tk.END)
        self.entry_meta_atual.insert(0, valores[2])
        self.entry_meta_data_inicial.delete(0, tk.END)
        self.entry_meta_data_inicial.insert(0, valores[4])
        self.entry_meta_data_final.delete(0, tk.END)
        self.entry_meta_data_final.insert(0, valores[5])
        self.meta_selecionada_id = meta_id

    def atualizar_meta(self):
        if not hasattr(self, "meta_selecionada_id") or not self.meta_selecionada_id:
            messagebox.showwarning("Aviso", "Selecione uma meta para atualizar.")
            return
        meta_id = self.meta_selecionada_id
        nome = self.entry_meta_nome.get().strip()
        valor_meta_texto = self.entry_meta_valor.get().strip().replace(",", ".")
        valor_atual_texto = self.entry_meta_atual.get().strip().replace(",", ".")
        data_inicial = self.entry_meta_data_inicial.get().strip()
        data_final = self.entry_meta_data_final.get().strip()
        if not nome or not valor_meta_texto or not data_inicial or not data_final:
            messagebox.showerror("Erro", "Preencha os campos obrigatórios da meta (Nome, Valor Meta, Data Inicial e Data Final).")
            return
        try:
            valor_meta = float(valor_meta_texto)
        except ValueError:
            messagebox.showerror("Erro", "O campo Valor Meta deve ser numérico.")
            return
        try:
            valor_atual = float(valor_atual_texto) if valor_atual_texto else 0.0
        except ValueError:
            messagebox.showerror("Erro", "O campo Valor Atual deve ser numérico.")
            return
        try:
            datetime.strptime(data_inicial, "%d/%m/%Y")
            datetime.strptime(data_final, "%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Erro", "Data Inicial ou Data Final inválida. Use o formato dd/mm/aaaa.")
            return
        try:
            self.cursor.execute(
                "UPDATE metas SET nome=?, valor_meta=?, valor_atual=?, data_inicial=?, data_final=? WHERE id=?",
                (nome, valor_meta, valor_atual, data_inicial, data_final, self.meta_selecionada_id)
            )
            self.conn.commit()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao atualizar meta no banco: {e}")
            return
        self.cursor.execute("SELECT id, nome, valor_meta, valor_atual, data_inicial, data_final FROM metas")
        self.metas = self.cursor.fetchall()
        self.atualizar_tabela_metas()
        messagebox.showinfo("Sucesso", "Meta atualizada com sucesso!")
        self.limpar_form_meta()

    def atualizar_tabela_metas(self):
        self.tabela_metas.delete(*self.tabela_metas.get_children())
        for meta in self.metas:
            meta_id, nome, valor_meta, valor_atual, data_inicial, data_final = meta
            percentual = (valor_atual / valor_meta * 100) if valor_meta > 0 else 0
            percentual_str = f"{percentual:.2f}%"
            tag = "meta_ok" if percentual >= 100 else "meta_incompleta"
            self.tabela_metas.insert("", "end", iid=str(meta_id), values=(nome, f"{valor_meta:.2f}".replace(".", ","), 
                                                                           f"{valor_atual:.2f}".replace(".", ","), percentual_str,
                                                                           data_inicial, data_final), tags=(tag,))
        self.tabela_metas.tag_configure("meta_ok", background="lightgreen")
        self.tabela_metas.tag_configure("meta_incompleta", background="lightcoral")

    def limpar_form_meta(self):
        self.entry_meta_nome.delete(0, tk.END)
        self.entry_meta_valor.delete(0, tk.END)
        self.entry_meta_atual.delete(0, tk.END)
        self.entry_meta_data_inicial.delete(0, tk.END)
        self.entry_meta_data_final.delete(0, tk.END)
        self.entry_meta_data_inicial.insert(0, "dd/mm/aaaa")
        self.entry_meta_data_final.insert(0, "dd/mm/aaaa")
        self.meta_selecionada_id = None

    # -------------------------
    # Aba Importação de Planilha
    # -------------------------
    def importar_planilha(self):
        filepath = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if not filepath:
            return
        try:
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            imported_data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    imported_data.append(row)
            if imported_data:
                self.despesas.extend(imported_data)
                self.atualizar_tabela()
                self.atualizar_indicador_gastos()
                self.atualizar_tabela_despesas_cartao()
                messagebox.showinfo("Sucesso", f"Importados {len(imported_data)} despesas da planilha.")
            else:
                messagebox.showinfo("Aviso", "Nenhuma despesa encontrada na planilha.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao importar planilha: {e}")

    # -------------------------
    # Aba Backup e Restauração
    # -------------------------
    def backup_db(self):
        backup_path = filedialog.asksaveasfilename(defaultextension=".db", filetypes=[("Database Files", "*.db")])
        if backup_path:
            try:
                self.conn.commit()
                shutil.copy("financeiro.db", backup_path)
                messagebox.showinfo("Backup", f"Backup realizado com sucesso em {backup_path}")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro no backup: {e}")

    def restaurar_backup(self):
        backup_path = filedialog.askopenfilename(defaultextension=".db", filetypes=[("Database Files", "*.db")])
        if backup_path:
            try:
                self.conn.commit()
                self.conn.close()
                shutil.copy(backup_path, "financeiro.db")
                self.conectar_banco()
                self.carregar_dados()
                self.atualizar_tabela()
                self.atualizar_tabela_despesas_cartao()
                self.atualizar_tabela_cartao()
                self.cursor.execute("SELECT id, nome, valor_meta, valor_atual, data_inicial, data_final FROM metas")
                self.metas = self.cursor.fetchall()
                self.atualizar_tabela_metas()
                self.combo_dashboard["values"] = [c[1] for c in self.cartoes]
                if self.cartoes:
                    self.combo_dashboard.current(0)
                    self.atualizar_dashboard_cartao()
                messagebox.showinfo("Backup", "Backup restaurado com sucesso!")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao restaurar backup: {e}")

if __name__ == "__main__":
    app = SpendingTracker()
    app.mainloop()
